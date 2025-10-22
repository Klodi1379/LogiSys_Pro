"""
Import/Export utilities for CSV and Excel files
"""
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO, StringIO
from typing import List, Dict, Any, Type
from django.db import models, transaction
from django.core.exceptions import ValidationError
import logging

logger = logging.getLogger(__name__)


class ImportExportService:
    """Service for importing and exporting data"""

    @staticmethod
    def export_to_csv(queryset, fields: List[str] = None, filename: str = 'export.csv') -> BytesIO:
        """
        Export queryset to CSV

        Args:
            queryset: Django queryset
            fields: List of field names to export (None for all)
            filename: Name for the export file

        Returns:
            BytesIO buffer containing CSV data
        """
        buffer = StringIO()
        writer = csv.writer(buffer)

        # Get model fields
        model = queryset.model
        if fields is None:
            fields = [f.name for f in model._meta.fields]

        # Write header
        headers = [model._meta.get_field(f).verbose_name.title() for f in fields]
        writer.writerow(headers)

        # Write data
        for obj in queryset:
            row = []
            for field in fields:
                value = getattr(obj, field)
                # Handle foreign keys
                if isinstance(value, models.Model):
                    value = str(value)
                row.append(value)
            writer.writerow(row)

        logger.info(f"Exported {queryset.count()} records to CSV")

        # Convert to BytesIO
        byte_buffer = BytesIO()
        byte_buffer.write(buffer.getvalue().encode('utf-8'))
        byte_buffer.seek(0)
        return byte_buffer

    @staticmethod
    def export_to_excel(queryset, fields: List[str] = None, filename: str = 'export.xlsx') -> BytesIO:
        """
        Export queryset to Excel with formatting

        Args:
            queryset: Django queryset
            fields: List of field names to export
            filename: Name for the export file

        Returns:
            BytesIO buffer containing Excel data
        """
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = queryset.model._meta.verbose_name_plural[:31]  # Excel sheet name limit

        # Get model fields
        model = queryset.model
        if fields is None:
            fields = [f.name for f in model._meta.fields]

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        headers = [model._meta.get_field(f).verbose_name.title() for f in fields]
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field in enumerate(fields, 1):
                value = getattr(obj, field)
                # Handle foreign keys
                if isinstance(value, models.Model):
                    value = str(value)
                # Handle dates/times
                elif hasattr(value, 'isoformat'):
                    value = value.isoformat()
                worksheet.cell(row=row_num, column=col_num, value=value)

        # Autosize columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save to buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        logger.info(f"Exported {queryset.count()} records to Excel")
        return buffer

    @staticmethod
    def import_from_csv(file, model: Type[models.Model], field_mapping: Dict[str, str]) -> tuple:
        """
        Import data from CSV file

        Args:
            file: File object or path
            model: Django model class
            field_mapping: Dict mapping CSV column names to model field names

        Returns:
            Tuple of (success_count, error_list)
        """
        success_count = 0
        errors = []

        try:
            # Read CSV
            if isinstance(file, str):
                with open(file, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
            else:
                file.seek(0)
                content = file.read().decode('utf-8')
                reader = csv.DictReader(StringIO(content))
                rows = list(reader)

            with transaction.atomic():
                for row_num, row in enumerate(rows, 2):  # Start from 2 (after header)
                    try:
                        # Map CSV columns to model fields
                        data = {}
                        for csv_col, model_field in field_mapping.items():
                            if csv_col in row:
                                data[model_field] = row[csv_col]

                        # Create object
                        obj = model(**data)
                        obj.full_clean()  # Validate
                        obj.save()
                        success_count += 1

                    except (ValidationError, Exception) as e:
                        errors.append({
                            'row': row_num,
                            'data': row,
                            'error': str(e)
                        })

        except Exception as e:
            logger.error(f"CSV import failed: {str(e)}")
            errors.append({'error': f"File processing error: {str(e)}"})

        logger.info(f"CSV import completed: {success_count} success, {len(errors)} errors")
        return success_count, errors

    @staticmethod
    def import_from_excel(file, model: Type[models.Model], field_mapping: Dict[str, str], sheet_name: str = None) -> tuple:
        """
        Import data from Excel file

        Args:
            file: File object or path
            model: Django model class
            field_mapping: Dict mapping Excel column names to model field names
            sheet_name: Name of sheet to import (None for first sheet)

        Returns:
            Tuple of (success_count, error_list)
        """
        success_count = 0
        errors = []

        try:
            # Load workbook
            if isinstance(file, str):
                workbook = openpyxl.load_workbook(file)
            else:
                file.seek(0)
                workbook = openpyxl.load_workbook(BytesIO(file.read()))

            # Get worksheet
            if sheet_name:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active

            # Get headers
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value)

            with transaction.atomic():
                for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
                    try:
                        # Create dict from row
                        row_data = dict(zip(headers, row))

                        # Map to model fields
                        data = {}
                        for excel_col, model_field in field_mapping.items():
                            if excel_col in row_data:
                                data[model_field] = row_data[excel_col]

                        # Create object
                        obj = model(**data)
                        obj.full_clean()
                        obj.save()
                        success_count += 1

                    except (ValidationError, Exception) as e:
                        errors.append({
                            'row': row_num,
                            'data': dict(zip(headers, row)),
                            'error': str(e)
                        })

        except Exception as e:
            logger.error(f"Excel import failed: {str(e)}")
            errors.append({'error': f"File processing error: {str(e)}"})

        logger.info(f"Excel import completed: {success_count} success, {len(errors)} errors")
        return success_count, errors

    @staticmethod
    def generate_template(model: Type[models.Model], format: str = 'csv', include_sample: bool = False) -> BytesIO:
        """
        Generate import template file

        Args:
            model: Django model class
            format: 'csv' or 'excel'
            include_sample: Include sample row

        Returns:
            BytesIO buffer containing template
        """
        fields = [f for f in model._meta.fields if not f.auto_created]
        field_names = [f.name for f in fields]
        headers = [f.verbose_name.title() for f in fields]

        if format == 'csv':
            buffer = StringIO()
            writer = csv.writer(buffer)
            writer.writerow(headers)

            if include_sample:
                sample_row = [''] * len(headers)
                writer.writerow(sample_row)

            byte_buffer = BytesIO()
            byte_buffer.write(buffer.getvalue().encode('utf-8'))
            byte_buffer.seek(0)
            return byte_buffer

        else:  # Excel
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Template"

            # Styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")

            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill

            # Add sample row if requested
            if include_sample:
                for col_num in range(1, len(headers) + 1):
                    worksheet.cell(row=2, column=col_num, value='')

            # Autosize
            for column in worksheet.columns:
                worksheet.column_dimensions[column[0].column_letter].width = 20

            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            return buffer
